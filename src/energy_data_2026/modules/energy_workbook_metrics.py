from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Mapping, Sequence

from openpyxl import load_workbook

from energy_data_2026.modules.energy_conversion import EnergyConversionCalculator


DEFAULT_ENERGY_ALIASES: Mapping[str, str] = {
    "电力": "电",
    "天然气": "燃气",
    "自来水": "自来水",
    "中水": "中水",
    "采暖热量": "采暖用热",
    "生活热水热量": "生活热水用热",
}


@dataclass(frozen=True)
class EnergyTypeMetrics:
    energy_type: str
    usage: float
    cost_yuan: float
    standard_coal_tce: float
    co2_ton: float


@dataclass(frozen=True)
class PeriodMetrics:
    year: int
    months: tuple[int, ...]
    total_cost_yuan: float
    total_standard_coal_tce: float
    total_co2_ton: float
    energy_metrics: tuple[EnergyTypeMetrics, ...]


@dataclass(frozen=True)
class MetricComparison:
    metric: str
    current_value: float
    previous_value: float
    change_value: float
    change_rate: float | None


def read_period_metrics(
    workbook_path: str | Path,
    year: int,
    months: Sequence[int],
    calculator: EnergyConversionCalculator,
    energy_aliases: Mapping[str, str] | None = None,
) -> PeriodMetrics:
    aliases = energy_aliases or DEFAULT_ENERGY_ALIASES
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook.active
        header_groups = _header_groups(worksheet)
        month_rows = _month_rows(worksheet)
        selected_rows = [_required_month_row(month_rows, month) for month in months]

        metrics: list[EnergyTypeMetrics] = []
        for workbook_energy_type, config_energy_type in aliases.items():
            columns = _energy_columns(worksheet, header_groups, workbook_energy_type)
            usage = sum(_to_float(worksheet.cell(row=row, column=columns["usage"]).value) for row in selected_rows)
            cost = sum(_to_float(worksheet.cell(row=row, column=columns["cost"]).value) for row in selected_rows)
            metrics.append(
                EnergyTypeMetrics(
                    energy_type=config_energy_type,
                    usage=round(usage, 4),
                    cost_yuan=round(cost, 2),
                    standard_coal_tce=calculator.calc_standard_coal_tce(config_energy_type, usage),
                    co2_ton=calculator.calc_co2_ton(config_energy_type, usage),
                )
            )

        return PeriodMetrics(
            year=year,
            months=tuple(months),
            total_cost_yuan=round(sum(item.cost_yuan for item in metrics), 2),
            total_standard_coal_tce=round(sum(item.standard_coal_tce for item in metrics), 4),
            total_co2_ton=round(sum(item.co2_ton for item in metrics), 4),
            energy_metrics=tuple(metrics),
        )
    finally:
        workbook.close()


def compare_period_metrics(current: PeriodMetrics, previous: PeriodMetrics) -> tuple[MetricComparison, ...]:
    return (
        _compare("能源成本(元)", current.total_cost_yuan, previous.total_cost_yuan),
        _compare("综合能耗(吨标准煤)", current.total_standard_coal_tce, previous.total_standard_coal_tce),
        _compare("二氧化碳排放(吨CO2)", current.total_co2_ton, previous.total_co2_ton),
    )


def subtract_period_metrics(total: PeriodMetrics, deducted: PeriodMetrics) -> PeriodMetrics:
    if total.year != deducted.year:
        raise ValueError(f"Cannot subtract different years: {total.year} and {deducted.year}")
    if total.months != deducted.months:
        raise ValueError(f"Cannot subtract different month ranges: {total.months} and {deducted.months}")

    deducted_by_type = {item.energy_type: item for item in deducted.energy_metrics}
    metrics: list[EnergyTypeMetrics] = []
    for total_item in total.energy_metrics:
        deducted_item = deducted_by_type.get(total_item.energy_type)
        if deducted_item is None:
            deducted_item = EnergyTypeMetrics(
                energy_type=total_item.energy_type,
                usage=0,
                cost_yuan=0,
                standard_coal_tce=0,
                co2_ton=0,
            )
        metrics.append(
            EnergyTypeMetrics(
                energy_type=total_item.energy_type,
                usage=round(total_item.usage - deducted_item.usage, 4),
                cost_yuan=round(total_item.cost_yuan - deducted_item.cost_yuan, 2),
                standard_coal_tce=round(total_item.standard_coal_tce - deducted_item.standard_coal_tce, 4),
                co2_ton=round(total_item.co2_ton - deducted_item.co2_ton, 4),
            )
        )

    return PeriodMetrics(
        year=total.year,
        months=total.months,
        total_cost_yuan=round(total.total_cost_yuan - deducted.total_cost_yuan, 2),
        total_standard_coal_tce=round(total.total_standard_coal_tce - deducted.total_standard_coal_tce, 4),
        total_co2_ton=round(total.total_co2_ton - deducted.total_co2_ton, 4),
        energy_metrics=tuple(metrics),
    )


def _compare(metric: str, current_value: float, previous_value: float) -> MetricComparison:
    change = round(current_value - previous_value, 4)
    rate = None if previous_value == 0 else round(change / previous_value, 4)
    return MetricComparison(
        metric=metric,
        current_value=current_value,
        previous_value=previous_value,
        change_value=change,
        change_rate=rate,
    )


def _header_groups(worksheet) -> dict[int, str]:
    groups: dict[int, str] = {}
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=column).value
        if value not in (None, ""):
            groups[column] = str(value).strip()

    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row <= 1 <= merged_range.max_row:
            value = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            if value in (None, ""):
                continue
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                groups[column] = str(value).strip()
    return groups


def _month_rows(worksheet) -> dict[int, int]:
    rows: dict[int, int] = {}
    for row in range(1, worksheet.max_row + 1):
        month = _month_number(worksheet.cell(row=row, column=1).value)
        if month is not None:
            rows[month] = row
    return rows


def _required_month_row(month_rows: Mapping[int, int], month: int) -> int:
    if month not in month_rows:
        raise ValueError(f"Month row not found: {month}")
    return month_rows[month]


def _energy_columns(worksheet, header_groups: Mapping[int, str], workbook_energy_type: str) -> dict[str, int]:
    usage_column: int | None = None
    cost_column: int | None = None
    for column in range(1, worksheet.max_column + 1):
        if header_groups.get(column) != workbook_energy_type:
            continue
        subheader = str(worksheet.cell(row=2, column=column).value or "")
        if usage_column is None and "实物量" in subheader:
            usage_column = column
        if cost_column is None and "成本" in subheader:
            cost_column = column

    if usage_column is None or cost_column is None:
        raise ValueError(f"Usage/cost columns not found for energy type: {workbook_energy_type}")
    return {"usage": usage_column, "cost": cost_column}


def _month_number(value: object) -> int | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if text.endswith("月"):
        text = text[:-1]
    if text.isdigit():
        month = int(text)
        if 1 <= month <= 12:
            return month
    return None


def _to_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        return 0.0
    return float(str(value).replace(",", "").strip())
