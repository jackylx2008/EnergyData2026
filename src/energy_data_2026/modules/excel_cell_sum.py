from __future__ import annotations

from dataclasses import dataclass
from numbers import Number
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


CellKey = tuple[int, int]


@dataclass(frozen=True)
class SheetSumResult:
    sheet_name: str
    written_cells: int
    skipped_merged_cells: int


@dataclass(frozen=True)
class WorkbookSumResult:
    target_file: Path
    sheet_results: tuple[SheetSumResult, ...]

    @property
    def written_cells(self) -> int:
        return sum(item.written_cells for item in self.sheet_results)

    @property
    def skipped_merged_cells(self) -> int:
        return sum(item.skipped_merged_cells for item in self.sheet_results)


@dataclass(frozen=True)
class SourceSheet:
    file: Path
    sheet_name: str


HeaderKey = tuple[str, str]


def sum_workbook_numeric_cells(
    first_source: str | Path,
    second_source: str | Path,
    target_file: str | Path,
    sheet_names: Iterable[str] | None = None,
) -> WorkbookSumResult:
    return sum_multiple_workbook_numeric_cells(
        source_files=(first_source, second_source),
        target_file=target_file,
        sheet_names=sheet_names,
    )


def sum_multiple_workbook_numeric_cells(
    source_files: Iterable[str | Path],
    target_file: str | Path,
    sheet_names: Iterable[str] | None = None,
) -> WorkbookSumResult:
    source_paths = tuple(Path(source_file) for source_file in source_files)
    if not source_paths:
        raise ValueError("At least one source Excel file is required")

    target_path = Path(target_file)
    _require_existing_files(*source_paths, target_path)

    source_workbooks = tuple(load_workbook(path, data_only=True, read_only=True) for path in source_paths)
    target_wb = load_workbook(target_path)

    try:
        names = tuple(sheet_names) if sheet_names else _common_sheet_names(*source_workbooks, target_wb)
        results: list[SheetSumResult] = []

        for sheet_name in names:
            for workbook, path in zip(source_workbooks, source_paths):
                _require_sheet(workbook, sheet_name, path)
            _require_sheet(target_wb, sheet_name, target_path)

            source_values = tuple(_numeric_cells(workbook[sheet_name]) for workbook in source_workbooks)
            cell_keys = set().union(*(values.keys() for values in source_values))
            written = 0
            skipped_merged = 0

            for row, column in sorted(cell_keys):
                target_cell = target_wb[sheet_name].cell(row=row, column=column)
                if isinstance(target_cell, MergedCell):
                    skipped_merged += 1
                    continue
                target_cell.value = sum(values.get((row, column), 0) for values in source_values)
                written += 1

            results.append(
                SheetSumResult(
                    sheet_name=sheet_name,
                    written_cells=written,
                    skipped_merged_cells=skipped_merged,
                )
            )

        target_wb.save(target_path)
        return WorkbookSumResult(target_file=target_path, sheet_results=tuple(results))
    finally:
        for workbook in source_workbooks:
            workbook.close()
        target_wb.close()


def copy_source_sheet_table_into_target_sheet(
    source_sheet: SourceSheet,
    target_file: str | Path,
    target_sheet_name: str,
    target_year: str = "2026年",
) -> WorkbookSumResult:
    target_path = Path(target_file)
    _require_existing_files(source_sheet.file, target_path)

    source_wb = load_workbook(source_sheet.file, data_only=True)
    target_wb = load_workbook(target_path)

    try:
        _require_sheet(source_wb, source_sheet.sheet_name, source_sheet.file)
        _require_sheet(target_wb, target_sheet_name, target_path)

        source_ws = source_wb[source_sheet.sheet_name]
        target_ws = target_wb[target_sheet_name]
        source_columns = _table_columns(source_ws, category_row=1, metric_row=2)
        target_columns = _table_columns(target_ws, category_row=3, metric_row=4)
        source_month_rows = _month_rows(source_ws, month_column=1)
        target_month_rows = _target_month_rows(target_ws, year_column=1, month_column=2, target_year=target_year)

        written = 0
        skipped_merged = 0
        for key, source_column in source_columns.items():
            target_column = target_columns.get(key)
            if target_column is None:
                continue

            for month, source_row in source_month_rows.items():
                target_row = target_month_rows.get(month)
                if target_row is None:
                    continue

                value = source_ws.cell(row=source_row, column=source_column).value
                if not _is_plain_number(value):
                    continue

                target_cell = target_ws.cell(row=target_row, column=target_column)
                if isinstance(target_cell, MergedCell):
                    skipped_merged += 1
                    continue
                target_cell.value = value
                written += 1

        target_wb.save(target_path)
        return WorkbookSumResult(
            target_file=target_path,
            sheet_results=(
                SheetSumResult(
                    sheet_name=target_sheet_name,
                    written_cells=written,
                    skipped_merged_cells=skipped_merged,
                ),
            ),
        )
    finally:
        source_wb.close()
        target_wb.close()


def sum_source_sheets_into_target_sheet(
    source_sheets: Iterable[SourceSheet],
    target_file: str | Path,
    target_sheet_name: str,
    target_year: str = "2026年",
) -> WorkbookSumResult:
    sources = tuple(source_sheets)
    if not sources:
        raise ValueError("At least one source Excel sheet is required")

    target_path = Path(target_file)
    _require_existing_files(*(source.file for source in sources), target_path)

    source_workbooks = tuple(load_workbook(source.file, data_only=True) for source in sources)
    target_wb = load_workbook(target_path)

    try:
        for source, workbook in zip(sources, source_workbooks):
            _require_sheet(workbook, source.sheet_name, source.file)
        _require_sheet(target_wb, target_sheet_name, target_path)

        target_sheet = target_wb[target_sheet_name]
        target_columns = _table_columns(target_sheet, category_row=3, metric_row=4)
        target_month_rows = _target_month_rows(
            target_sheet,
            year_column=1,
            month_column=2,
            target_year=target_year,
        )
        source_tables = tuple(
            (
                workbook[source.sheet_name],
                _table_columns(workbook[source.sheet_name], category_row=1, metric_row=2),
                _month_rows(workbook[source.sheet_name], month_column=1),
            )
            for source, workbook in zip(sources, source_workbooks)
        )

        if target_columns and target_month_rows and all(columns and rows for _, columns, rows in source_tables):
            written, skipped_merged = _write_mapped_table_values(
                target_sheet=target_sheet,
                target_columns=target_columns,
                target_month_rows=target_month_rows,
                source_tables=source_tables,
            )
        else:
            source_values = tuple(
                _numeric_cells(workbook[source.sheet_name])
                for source, workbook in zip(sources, source_workbooks)
            )
            written, skipped_merged = _write_coordinate_sum_values(
                target_sheet=target_sheet,
                source_values=source_values,
            )

        target_wb.save(target_path)
        return WorkbookSumResult(
            target_file=target_path,
            sheet_results=(
                SheetSumResult(
                    sheet_name=target_sheet_name,
                    written_cells=written,
                    skipped_merged_cells=skipped_merged,
                ),
            ),
        )
    finally:
        for workbook in source_workbooks:
            workbook.close()
        target_wb.close()


def _require_existing_files(*paths: Path) -> None:
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise FileNotFoundError("Excel file not found: " + "; ".join(missing))


def _require_sheet(workbook, sheet_name: str, path: Path) -> None:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in {path}")


def _common_sheet_names(*workbooks) -> tuple[str, ...]:
    common = set(workbooks[0].sheetnames)
    for workbook in workbooks[1:]:
        common &= set(workbook.sheetnames)
    return tuple(sheet_name for sheet_name in workbooks[-1].sheetnames if sheet_name in common)


def _numeric_cells(worksheet) -> dict[CellKey, Number]:
    values: dict[CellKey, Number] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if _is_plain_number(cell.value):
                values[(cell.row, cell.column)] = cell.value
    return values


def _is_plain_number(value: object) -> bool:
    return isinstance(value, Number) and not isinstance(value, bool)


def _write_mapped_table_values(
    target_sheet,
    target_columns: dict[HeaderKey, int],
    target_month_rows: dict[str, int],
    source_tables,
) -> tuple[int, int]:
    written = 0
    skipped_merged = 0
    for key, target_column in target_columns.items():
        for month, target_row in target_month_rows.items():
            total = 0
            has_value = False
            for source_sheet, source_columns, source_month_rows in source_tables:
                source_column = source_columns.get(key)
                source_row = source_month_rows.get(month)
                if source_column is None or source_row is None:
                    continue
                value = source_sheet.cell(row=source_row, column=source_column).value
                if not _is_plain_number(value):
                    continue
                total += value
                has_value = True
            if not has_value:
                continue

            target_cell = target_sheet.cell(row=target_row, column=target_column)
            if isinstance(target_cell, MergedCell):
                skipped_merged += 1
                continue
            target_cell.value = total
            written += 1
    return written, skipped_merged


def _write_coordinate_sum_values(target_sheet, source_values: tuple[dict[CellKey, Number], ...]) -> tuple[int, int]:
    cell_keys = set().union(*(values.keys() for values in source_values))
    written = 0
    skipped_merged = 0
    for row, column in sorted(cell_keys):
        target_cell = target_sheet.cell(row=row, column=column)
        if isinstance(target_cell, MergedCell):
            skipped_merged += 1
            continue
        target_cell.value = sum(values.get((row, column), 0) for values in source_values)
        written += 1
    return written, skipped_merged


def _table_columns(worksheet, category_row: int, metric_row: int) -> dict[HeaderKey, int]:
    columns: dict[HeaderKey, int] = {}
    for column in range(1, worksheet.max_column + 1):
        category = _normalized_text(_merged_cell_value(worksheet, category_row, column))
        metric = _normalized_text(_merged_cell_value(worksheet, metric_row, column))
        if not category or not metric:
            continue
        columns[(category, metric)] = column
    return columns


def _month_rows(worksheet, month_column: int) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row in range(1, worksheet.max_row + 1):
        month = _normalized_month(_merged_cell_value(worksheet, row, month_column))
        if month:
            rows[month] = row
    return rows


def _target_month_rows(worksheet, year_column: int, month_column: int, target_year: str) -> dict[str, int]:
    rows: dict[str, int] = {}
    normalized_year = _normalized_text(target_year)
    for row in range(1, worksheet.max_row + 1):
        month = _normalized_month(_merged_cell_value(worksheet, row, month_column))
        if not month:
            continue
        year = _nearest_year_value(worksheet, row=row, year_column=year_column)
        if _normalized_text(year) == normalized_year:
            rows[month] = row
    return rows


def _nearest_year_value(worksheet, row: int, year_column: int) -> object:
    for current_row in range(row, 0, -1):
        value = _merged_cell_value(worksheet, current_row, year_column)
        if value:
            return value
    return None


def _merged_cell_value(worksheet, row: int, column: int) -> object:
    cell = worksheet.cell(row=row, column=column)
    if not isinstance(cell, MergedCell):
        return cell.value
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None


def _normalized_month(value: object) -> str | None:
    text = _normalized_text(value)
    if not text.endswith("月"):
        return None
    month_number = text[:-1]
    if not month_number.isdigit():
        return None
    month = int(month_number)
    if month < 1 or month > 12:
        return None
    return f"{month}月"


def _normalized_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", "").replace(" ", "").strip()
