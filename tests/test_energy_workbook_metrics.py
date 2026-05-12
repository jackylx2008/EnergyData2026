from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from energy_data_2026.modules.energy_conversion import EnergyConversionCalculator
from energy_data_2026.modules.energy_workbook_metrics import (
    EnergyTypeMetrics,
    PeriodMetrics,
    read_period_metrics,
    subtract_period_metrics,
)
from energy_data_2026.flows.quarter_compare import quarter_months


class EnergyWorkbookMetricsTests(unittest.TestCase):
    def test_quarter_months(self) -> None:
        self.assertEqual(quarter_months(1), (1, 2, 3))
        self.assertEqual(quarter_months(2), (4, 5, 6))
        self.assertEqual(quarter_months(3), (7, 8, 9))
        self.assertEqual(quarter_months(4), (10, 11, 12))

    def test_subtract_period_metrics(self) -> None:
        total = PeriodMetrics(
            year=2026,
            months=(1, 2, 3),
            total_cost_yuan=100,
            total_standard_coal_tce=10,
            total_co2_ton=20,
            energy_metrics=(
                EnergyTypeMetrics("电", 1000, 80, 8, 16),
                EnergyTypeMetrics("燃气", 10, 20, 2, 4),
            ),
        )
        deducted = PeriodMetrics(
            year=2026,
            months=(1, 2, 3),
            total_cost_yuan=40,
            total_standard_coal_tce=4,
            total_co2_ton=8,
            energy_metrics=(
                EnergyTypeMetrics("电", 400, 30, 3, 6),
                EnergyTypeMetrics("燃气", 5, 10, 1, 2),
            ),
        )

        result = subtract_period_metrics(total, deducted)

        self.assertEqual(result.total_cost_yuan, 60)
        self.assertEqual(result.total_standard_coal_tce, 6)
        self.assertEqual(result.total_co2_ton, 12)
        self.assertEqual(result.energy_metrics[0].usage, 600)
        self.assertEqual(result.energy_metrics[1].cost_yuan, 10)

    def test_reads_q1_metrics_from_merged_headers(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "energy.xlsx"
            self._make_workbook(path)
            calculator = EnergyConversionCalculator(
                coal_conversion={"电": 0.1229, "燃气": 1.33},
                co2_conversion={"电": 0.5, "燃气": 2.0},
            )

            result = read_period_metrics(
                path,
                year=2026,
                months=(1, 2, 3),
                calculator=calculator,
                energy_aliases={"电力": "电", "天然气": "燃气"},
            )

            self.assertEqual(result.total_cost_yuan, 660.0)
            self.assertEqual(result.total_standard_coal_tce, 0.1535)
            self.assertEqual(result.total_co2_ton, 0.42)

    @staticmethod
    def _make_workbook(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.merge_cells("A1:A2")
        sheet.merge_cells("B1:C1")
        sheet.merge_cells("D1:E1")
        sheet["A1"] = "月度"
        sheet["B1"] = "电力"
        sheet["D1"] = "天然气"
        sheet["B2"] = "实物量  （千瓦时）"
        sheet["C2"] = "成本（元）"
        sheet["D2"] = "实物量  （立方米）"
        sheet["E2"] = "成本（元）"
        rows = [
            ["1月", 100, 100, 10, 10],
            ["2月", 200, 200, 20, 20],
            ["3月", 300, 300, 30, 30],
        ]
        for index, row in enumerate(rows, start=3):
            for column, value in enumerate(row, start=1):
                sheet.cell(row=index, column=column).value = value
        workbook.save(path)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
