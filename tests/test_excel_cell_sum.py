from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from energy_data_2026.modules.excel_cell_sum import sum_workbook_numeric_cells


class ExcelCellSumTests(unittest.TestCase):
    def test_sums_numeric_cells_and_preserves_target_styles(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first = root / "first.xlsx"
            second = root / "second.xlsx"
            target = root / "target.xlsx"

            self._make_workbook(first, a1=1, b1=2, c1="skip")
            self._make_workbook(second, a1=3, b1=None, c1=7)
            self._make_target(target)

            result = sum_workbook_numeric_cells(first, second, target)

            self.assertEqual(result.written_cells, 3)
            workbook = load_workbook(target)
            sheet = workbook.active
            self.assertEqual(sheet["A1"].value, 4)
            self.assertEqual(sheet["B1"].value, 2)
            self.assertEqual(sheet["C1"].value, 7)
            self.assertEqual(sheet["A1"].fill.fgColor.rgb, "00FFFF00")
            self.assertTrue(sheet["A1"].font.bold)
            workbook.close()

    @staticmethod
    def _make_workbook(path: Path, a1: object, b1: object, c1: object) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet["A1"] = a1
        sheet["B1"] = b1
        sheet["C1"] = c1
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _make_target(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet["A1"] = 0
        sheet["A1"].font = Font(bold=True, color="00FF0000")
        sheet["A1"].fill = PatternFill("solid", fgColor="00FFFF00")
        sheet["B1"] = 0
        sheet["C1"] = 0
        workbook.save(path)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
