from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from energy_data_2026.context import AppContext
from energy_data_2026.flows.huitou_sum import run


class HuitouSumFlowTests(unittest.TestCase):
    def test_reads_excel_paths_from_yaml_config_shape(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first = root / "first.xlsx"
            second = root / "second.xlsx"
            target = root / "target.xlsx"

            self._make_workbook(first, value=1)
            self._make_workbook(second, value=2)
            self._make_workbook(target, value=0)

            result = run(
                AppContext(
                    project_root=root,
                    env={},
                    config={
                        "huitou_sum": {
                            "source_files": [str(first), str(second)],
                            "target_file": str(target),
                        }
                    },
                )
            )

            self.assertEqual(result.written_cells, 1)
            workbook = load_workbook(target)
            self.assertEqual(workbook.active["A1"].value, 3)
            workbook.close()

    def test_sums_different_sources_for_each_target_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            sheet1_first = root / "sheet1_first.xlsx"
            sheet1_second = root / "sheet1_second.xlsx"
            sheet2_first = root / "sheet2_first.xlsx"
            sheet2_second = root / "sheet2_second.xlsx"
            target = root / "target.xlsx"

            self._make_named_workbook(sheet1_first, "用量", value=1)
            self._make_named_workbook(sheet1_second, "用量", value=2)
            self._make_named_workbook(sheet2_first, "费用", value=10)
            self._make_named_workbook(sheet2_second, "费用", value=20)
            self._make_target_workbook(target, ("用量", "费用"))

            result = run(
                AppContext(
                    project_root=root,
                    env={},
                    config={
                        "huitou_sum": {
                            "target_file": str(target),
                            "jobs": [
                                {
                                    "sheet_name": "用量",
                                    "source_files": [str(sheet1_first), str(sheet1_second)],
                                },
                                {
                                    "sheet_name": "费用",
                                    "source_files": [str(sheet2_first), str(sheet2_second)],
                                },
                            ],
                        }
                    },
                )
            )

            self.assertEqual(result.written_cells, 2)
            workbook = load_workbook(target)
            self.assertEqual(workbook["用量"]["A1"].value, 3)
            self.assertEqual(workbook["费用"]["A1"].value, 30)
            workbook.close()

    def test_sums_source_sheets_into_different_target_sheet_name(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first = root / "first.xlsx"
            second = root / "second.xlsx"
            target = root / "target.xlsx"

            self._make_named_workbook(first, "SourceA", value=5)
            self._make_named_workbook(second, "SourceB", value=6)
            self._make_target_workbook(target, ("Target",))

            result = run(
                AppContext(
                    project_root=root,
                    env={},
                    config={
                        "huitou_sum": {
                            "target_file": str(target),
                            "jobs": [
                                {
                                    "sheet_name": "Target",
                                    "source_files": [
                                        {"file": str(first), "sheet_name": "SourceA"},
                                        {"file": str(second), "sheet_name": "SourceB"},
                                    ],
                                }
                            ],
                        }
                    },
                )
            )

            self.assertEqual(result.written_cells, 1)
            workbook = load_workbook(target)
            self.assertEqual(workbook["Target"]["A1"].value, 11)
            workbook.close()

    def test_copies_values_when_target_sheet_has_one_source_file(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source.xlsx"
            target = root / "target.xlsx"

            self._make_source_table_workbook(source, "Source")
            self._make_target_table_workbook(target, "Target")

            result = run(
                AppContext(
                    project_root=root,
                    env={},
                    config={
                        "huitou_sum": {
                            "target_file": str(target),
                            "jobs": [
                                {
                                    "sheet_name": "Target",
                                    "source_files": [
                                        {"file": str(source), "sheet_name": "Source"},
                                    ],
                                }
                            ],
                        }
                    },
                )
            )

            self.assertEqual(result.written_cells, 4)
            workbook = load_workbook(target)
            sheet = workbook["Target"]
            self.assertEqual(sheet["C5"].value, 100)
            self.assertEqual(sheet["D5"].value, 200)
            self.assertEqual(sheet["C6"].value, 300)
            self.assertEqual(sheet["D6"].value, 400)
            workbook.close()

    def test_reads_workbook_paths_from_env_keys(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source.xlsx"
            target = root / "target.xlsx"

            self._make_source_table_workbook(source, "Source")
            self._make_target_table_workbook(target, "Target")

            result = run(
                AppContext(
                    project_root=root,
                    env={
                        "HUITOU_TARGET_FILE": str(target),
                        "HUITOU_SOURCE_FILE": str(source),
                    },
                    config={
                        "huitou_sum": {
                            "target_file_key": "HUITOU_TARGET_FILE",
                            "jobs": [
                                {
                                    "sheet_name": "Target",
                                    "source_files": [
                                        {"file_key": "HUITOU_SOURCE_FILE", "sheet_name": "Source"},
                                    ],
                                }
                            ],
                        }
                    },
                )
            )

            self.assertEqual(result.written_cells, 4)
            workbook = load_workbook(target)
            self.assertEqual(workbook["Target"]["C5"].value, 100)
            workbook.close()

    def test_reads_sheet_names_from_env_keys(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source.xlsx"
            target = root / "target.xlsx"

            self._make_source_table_workbook(source, "Source")
            self._make_target_table_workbook(target, "Target")

            result = run(
                AppContext(
                    project_root=root,
                    env={
                        "HUITOU_TARGET_FILE": str(target),
                        "HUITOU_SOURCE_FILE": str(source),
                        "HUITOU_TARGET_SHEET": "Target",
                        "HUITOU_SOURCE_SHEET": "Source",
                    },
                    config={
                        "huitou_sum": {
                            "target_file_key": "HUITOU_TARGET_FILE",
                            "jobs": [
                                {
                                    "sheet_name_key": "HUITOU_TARGET_SHEET",
                                    "source_files": [
                                        {
                                            "file_key": "HUITOU_SOURCE_FILE",
                                            "sheet_name_key": "HUITOU_SOURCE_SHEET",
                                        },
                                    ],
                                }
                            ],
                        }
                    },
                )
            )

            self.assertEqual(result.written_cells, 4)
            workbook = load_workbook(target)
            self.assertEqual(workbook["Target"]["D6"].value, 400)
            workbook.close()

    @staticmethod
    def _make_workbook(path: Path, value: int) -> None:
        workbook = Workbook()
        workbook.active["A1"] = value
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _make_named_workbook(path: Path, sheet_name: str, value: int) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet["A1"] = value
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _make_target_workbook(path: Path, sheet_names: tuple[str, ...]) -> None:
        workbook = Workbook()
        first_sheet = workbook.active
        first_sheet.title = sheet_names[0]
        first_sheet["A1"] = 0
        for sheet_name in sheet_names[1:]:
            sheet = workbook.create_sheet(sheet_name)
            sheet["A1"] = 0
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _make_source_table_workbook(path: Path, sheet_name: str) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.merge_cells("B1:C1")
        sheet["B1"] = "电力"
        sheet["A2"] = "月度"
        sheet["B2"] = "实物量（千瓦时）"
        sheet["C2"] = "成本（元）"
        sheet["A3"] = "1月"
        sheet["B3"] = 100
        sheet["C3"] = 200
        sheet["A4"] = "2月"
        sheet["B4"] = 300
        sheet["C4"] = 400
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _make_target_table_workbook(path: Path, sheet_name: str) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.merge_cells("A5:A7")
        sheet["A5"] = "2026年"
        sheet["B5"] = "1月"
        sheet["B6"] = "2月"
        sheet["B7"] = "一季度"
        sheet.merge_cells("C3:D3")
        sheet["C3"] = "电力"
        sheet["C4"] = "实物量（千瓦时）"
        sheet["D4"] = "成本（元）"
        workbook.save(path)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
